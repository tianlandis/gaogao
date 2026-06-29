#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""舞蹈类数据构建：分段表 -> compressed_ranks.json, 投档计划 -> data.json"""
import os, json, sys
if sys.platform == 'win32':
    import io; sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
FDB = os.path.join(BASE, 'fen_duan_biao')

try:
    import xlrd
except:
    print("需要安装 xlrd: pip install xlrd")
    sys.exit(1)
try:
    import openpyxl
except:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ==================== Part 1: 分段表 -> compressed_ranks.json ====================

def parse_fdb_xls(filepath):
    """解析一分一段表 XLS/XLSX, 返回 {score_str: cumulative_count}"""
    result = {}
    
    # 先尝试 xlrd (兼容 .xls 和伪装成.xlsx的旧格式)
    try:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        rows = ws.nrows
        data_start = 0
        for i in range(min(10, rows)):
            vals = [str(ws.cell_value(i, c)).strip() for c in range(ws.ncols)]
            if '分数段' in ''.join(vals) or '成绩' in ''.join(vals):
                data_start = i + 1
                break
        if data_start == 0: data_start = 2
        
        for i in range(data_start, rows):
            try:
                score = ws.cell_value(i, 0)
                cum = ws.cell_value(i, 2)
                if score == '' or cum == '':
                    continue
                score_f = float(str(score).strip())
                cum_i = int(float(cum))
                result[f'{score_f:.2f}'] = cum_i
            except:
                continue
        if result:
            return result
    except:
        pass
    
    # 回退到 openpyxl (真正的 .xlsx)
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        rows = ws.max_row
        data_start = 0
        for i in range(1, min(10, rows + 1)):
            vals = [str(ws.cell(i, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            if '分数段' in ''.join(vals) or '成绩' in ''.join(vals):
                data_start = i + 1
                break
        if data_start == 0: data_start = 3
        
        for i in range(data_start, rows + 1):
            try:
                score = ws.cell(i, 1).value
                cum = ws.cell(i, 3).value
                if score is None or cum is None:
                    continue
                score_f = float(str(score).strip())
                cum_i = int(float(cum))
                result[f'{score_f:.2f}'] = cum_i
            except:
                continue
    except:
        pass
    
    return result


# 文件映射: key -> (filename, year)
FILE_MAP = {
    'dance_2024': ('2024_综合.xlsx', '2024'),
    'dance_2025': ('2025_综合.xls', '2025'),
    'dance_2026': ('2026_综合.xls', '2026'),
    'dance_art_2024': ('2024_专业_原始.xls', '2024'),
    'dance_art_2025': ('2025_专业_原始.xls', '2025'),
    'dance_art_2026': ('2026_专业_原始.xls', '2026'),
    'dance_shuang_art_2024': ('2024_双达线_专业.xlsx', '2024'),
    'dance_shuang_art_2025': ('2025_双达线_专业.xls', '2025'),
    'dance_shuang_art_2026': ('2026_双达线_专业.xls', '2026'),
    'dance_culture_2024': ('2024_文化.xlsx', '2024'),
    'dance_culture_2025': ('2025_文化.xls', '2025'),
}

compressed_ranks = {}
for key, (filename, year) in FILE_MAP.items():
    fp = os.path.join(FDB, year, filename)
    if not os.path.exists(fp):
        print(f'[WARN] {fp} 不存在')
        continue
    data = parse_fdb_xls(fp)
    compressed_ranks[key] = data
    print(f'[OK] {key}: {len(data)} 条 (file: {filename})')

# 保存
output_path = os.path.join(BASE, 'compressed_ranks.json')
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(compressed_ranks, f, ensure_ascii=False)
print(f'\nSaved compressed_ranks.json ({os.path.getsize(output_path)/1024:.1f} KB)')

# ==================== Part 2: 投档计划 -> data.json ====================

PLAN_FILE = os.path.join(BASE, '投档计划数据舞蹈类.xlsx')
wb = openpyxl.load_workbook(PLAN_FILE)
ws = wb.active

headers = [c.value for c in ws[1]]
print(f'\n投档计划: {ws.max_row - 1} data rows, {len(headers)} cols')

def safe_int(v):
    """安全转整数"""
    if v is None or v == '' or v == '免费' or v == '—':
        return 0
    try:
        return int(float(str(v).replace(',', '').replace(' ', '')))
    except:
        return 0

def safe_float(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(str(v).replace(',', '').replace(' ', ''))
    except:
        return 0.0

def safe_str(v):
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ('nan', 'None', '—') else s

schools = []
for row_idx in range(2, ws.max_row + 1):
    cells = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
    
    code = safe_str(cells[1])           # 院校代码
    name = safe_str(cells[2])           # 院校名称
    major_code = safe_str(cells[3])     # 专业代码
    major = safe_str(cells[4])          # 专业名称
    nature = safe_str(cells[5])         # 办学性质
    subject_req = safe_str(cells[6])    # 选科
    duration = safe_int(cells[7])       # 学制
    tuition = safe_str(cells[8])        # 学费（可能是"免费"）
    plan = safe_int(cells[9])           # 26人
    plan2025 = safe_int(cells[10])      # 25人
    # cells[11] = 25投 (提交人数, skip)
    admitScore = safe_float(cells[12])  # 25录分
    line = safe_float(cells[13])        # 25投分
    rank = safe_int(cells[14])          # 25投位
    code2025 = safe_str(cells[15])      # 25代码
    major_info2025 = safe_str(cells[16]) # 25专业信息
    plan2024 = safe_int(cells[17])       # 24人
    line2024 = safe_float(cells[18])     # 24综合分
    rank2024 = safe_int(cells[19])       # 24位
    province = safe_str(cells[20])       # 省份
    city = safe_str(cells[21])           # 城市
    cityRating = safe_str(cells[22])     # 城市评级
    cityRank = safe_int(cells[23])       # 城市排行
    # cells[24] = 主管部门 (skip for now, not in music data)
    schoolLevel = safe_str(cells[25])    # 院校水平
    schoolBg = safe_str(cells[26])       # 院校背景
    schoolNick = safe_str(cells[27])     # 院校花称
    majorEval = safe_str(cells[28])      # 专业评估
    hasMaster = safe_int(cells[29])      # 硕士点
    masterMajor = safe_str(cells[30])    # 硕士专业
    hasDoctor = safe_int(cells[31])      # 博士点
    doctorMajor = safe_str(cells[32])    # 博士专业
    majorLevel = safe_str(cells[33])     # 专业水平
    level = safe_str(cells[34])          # 本专科
    schoolRank = safe_int(cells[35])     # 院校排名
    softRank = safe_str(cells[36])       # 软科校排
    softMajorRank = safe_str(cells[37])  # 软科专业排名
    majorRank = safe_str(cells[38])      # 专业排名
    majorRankPct = safe_str(cells[39])   # 专业排名比例
    majorTotal = safe_str(cells[40])     # 专业总数
    admissionUrl = safe_str(cells[41])   # 招生章程
    schoolInfoUrl = safe_str(cells[42])  # 学校招生信息
    vrUrl = safe_str(cells[43])          # 校园VR
    baikeUrl = safe_str(cells[44])       # 院校百科
    employUrl = safe_str(cells[45])      # 就业质量
    
    # 计算缺失字段
    batch = '本科批' if '本科' in level else '专科批' if '专科' in level else '本科批'
    sourceBatch = '舞蹈类'
    
    # 学费处理
    tuition_val = 0
    if tuition == '免费':
        tuition_val = 0
    elif tuition.isdigit():
        tuition_val = int(tuition)
    else:
        try:
            tuition_val = int(float(tuition))
        except:
            tuition_val = 0
    
    rec = {
        'code': code,
        'name': name,
        'major_code': major_code,
        'major': major,
        'sourceBatch': sourceBatch,
        'direction': 'dance',
        'province': province,
        'city': city,
        'level': level,
        'nature': nature,
        'batch': batch,
        'subject_req': subject_req,
        'duration': duration,
        'tuition': tuition_val,
        'plan': plan,
        'line': line,
        'rank': rank,
        'plan2025': plan2025,
        'code2025': code2025,
        'major_code2025': major_info2025,  # 存储原专业信息
        'admitScore': admitScore,
        'line2024': line2024,
        'rank2024': rank2024,
        'plan2024': plan2024,
        'cityRating': cityRating,
        'cityRank': cityRank,
        'schoolLevel': schoolLevel,
        'schoolBg': schoolBg,
        'schoolNick': schoolNick,
        'majorEval': majorEval,
        'hasMaster': hasMaster,
        'masterMajor': masterMajor,
        'hasDoctor': hasDoctor,
        'doctorMajor': doctorMajor,
        'majorLevel': majorLevel,
        'schoolRank': schoolRank,
        'softRank': softRank,
        'softMajorRank': softMajorRank,
        'majorRank': majorRank,
        'majorRankPct': majorRankPct,
        'majorTotal': majorTotal,
        'admissionUrl': admissionUrl,
        'schoolInfoUrl': schoolInfoUrl,
        'vrUrl': vrUrl,
        'baikeUrl': baikeUrl,
        'employUrl': employUrl,
    }
    schools.append(rec)

# 保存 data.json
data_output = {
    'metadata': {
        'version': '1.0',
        'update_time': '2026-06-27',
        'source': '投档计划数据舞蹈类.xlsx',
        'year': 2026,
        'category': '艺术统考舞蹈类',
        'description': '2024-2026山东省舞蹈类统一招生计划与投档数据（含院校标签）'
    },
    'schools': schools
}
data_path = os.path.join(BASE, 'data.json')
with open(data_path, 'w', encoding='utf-8') as f:
    json.dump(data_output, f, ensure_ascii=False, indent=2)
print(f'Saved data.json: {len(schools)} schools ({os.path.getsize(data_path)/1024:.1f} KB)')

# 统计
print(f'\n=== 统计 ===')
print(f'院校数: {len(set(s["code"] for s in schools))}')
print(f'专业数: {len(schools)}')
print(f'省份分布: {len(set(s["province"] for s in schools))} 省')
for prov in sorted(set(s['province'] for s in schools)):
    cnt = sum(1 for s in schools if s['province'] == prov)
    print(f'  {prov}: {cnt}')
print(f'有2025投档线: {sum(1 for s in schools if s["line"] > 0)}')
print(f'有2024投档线: {sum(1 for s in schools if s["line2024"] > 0)}')
print(f'2026计划已知: {sum(1 for s in schools if s["plan"] > 0)}')
print(f'tuition为0: {sum(1 for s in schools if s["tuition"] == 0)}')

print('\nDONE!')

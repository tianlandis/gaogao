#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从Excel读取2026年音乐类招生计划 -> plan_2026.json"""

import json, sys, os, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

from collections import defaultdict

EXCEL_PATH = r'E:\gaogao\downloads\2026音乐生院校招生计划表（不完整）.xlsx'

# ======== 配置 ========
# 钢琴/键盘关键词（器乐方向只保留这些）
PIANO_KW = ['钢琴', '键盘', '电子管风琴', '双排键']

# 类别映射
CATEGORY_MAP = {
    '音乐类(音乐表演-声乐)': 'vocal',
    '音乐类(音乐表演-器乐)': 'instrumental',
    '音乐类(音乐教育)': 'education',
}

# 提前批院校代码（手动指定，因为Excel没有批次列）
ADVANCE_BATCH_CODES = {
    'A027',  # 北京师范大学
    'A033',  # 中国传媒大学
    'A045',  # 中央音乐学院
    'A046',  # 中国音乐学院
    'A278',  # 上海音乐学院
    'A072',  # 天津音乐学院
    'A524',  # 武汉音乐学院
    'A177',  # 沈阳音乐学院
    'A587',  # 星海音乐学院
    'A654',  # 四川音乐学院
    'A728',  # 西安音乐学院
    'A535',  # 浙江音乐学院
    'A236',  # 哈尔滨音乐学院
    'A384',  # 厦门大学
}

def is_advance_batch(code, name):
    """判断是否为提前批"""
    if code in ADVANCE_BATCH_CODES:
        return True
    # 音乐学院/艺术学院一般提前批
    kw = ['音乐学院', '艺术学院', '中央戏剧', '北京电影', '中国戏曲', '上海戏剧']
    if any(k in name for k in kw):
        return True
    return False

def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None or row[1] is None or row[2] is None:
            continue
        rows.append(row)
    print(f"Excel读取: {len(rows)} 条记录 (Sheet: {ws.title})")
    return rows

def main():
    rows = load_excel(EXCEL_PATH)
    
    plans = []
    seen = set()
    stats = defaultdict(lambda: {'count': 0, 'total_plan': 0})
    skipped_instrumental = 0
    
    for row in rows:
        source_batch = str(row[0]).strip() if row[0] else ''          # 科类名称
        code = str(row[1]).strip() if row[1] else ''                   # 院校代码
        name = str(row[2]).strip() if row[2] else ''                   # 院校名称
        major_code = str(row[3]).strip() if row[3] else ''             # 专业代码
        major = str(row[4]).strip() if row[4] else ''                  # 专业名称
        nature = str(row[5]).strip() if row[5] else ''                 # 办学性质
        subject_req = str(row[6]).strip() if row[6] else ''            # 选科
        years = row[7] if row[7] else ''                               # 学制
        fee = row[8] if row[8] else ''                                 # 学费
        plan_num = row[9] if row[9] else 0                             # 计划
        province = str(row[10]).strip() if len(row) > 10 and row[10] else ''   # 省份
        city = str(row[11]).strip() if len(row) > 11 and row[11] else ''       # 城市
        
        # 映射方向
        direction = CATEGORY_MAP.get(source_batch, 'education')
        
        # 钢琴优先 remap
        if direction == 'instrumental':
            if any(kw in major for kw in ['钢琴', '键盘', '电子管风琴', '双排键']):
                pass  # keep instrumental
            else:
                skipped_instrumental += 1
                continue  # 跳过非钢琴器乐
        
        # 再次检查声乐方向的专业是否含器乐关键词(部分声乐方向也含钢琴)
        if direction == 'vocal' and any(kw in major for kw in PIANO_KW):
            # 声乐方向但专业是钢琴相关，改为instrumental
            direction = 'instrumental'
        
        # 计划数
        try:
            plan_num = int(plan_num) if plan_num else 0
        except:
            plan_num = 0
        
        # 批量
        batch = '本科提前批' if is_advance_batch(code, name) else '本科批'
        
        # 学费
        try:
            fee = int(fee) if fee else ''
        except:
            fee = str(fee) if fee else ''
        
        # 学制
        try:
            years = int(years) if years else ''
        except:
            years = str(years) if years else ''
        
        # 去重
        key = f"{code}_{name}_{major[:60]}_{direction}"
        if key in seen:
            continue
        seen.add(key)
        
        plans.append({
            'source_batch': source_batch,       # 科类名称
            'code': code,                       # 院校代码
            'name': name,                       # 院校名称
            'major_code': major_code,           # 专业代码
            'major': major,                     # 专业名称
            'nature': nature,                   # 办学性质
            'subject_req': subject_req,         # 选科
            'years': years,                     # 学制
            'fee': fee,                         # 学费
            'plan': plan_num,                   # 计划
            'province': province,               # 省份
            'city': city,                       # 城市
            'direction': direction,             # 内部方向
            'batch': batch,                     # 录取批次
            'remark': '',                       # 备注（从专业名提取要求）
        })
        
        stats[direction]['count'] += 1
        stats[direction]['total_plan'] += plan_num
    
    # 统计
    total_schools = len(set(p['name'] for p in plans))
    total_plans = sum(p['plan'] for p in plans)
    batch_counts = defaultdict(int)
    for p in plans:
        batch_counts[p['batch']] += 1
    
    print(f"\n=== 处理结果 ===")
    print(f"总记录数: {len(plans)}")
    print(f"院校数: {total_schools}")
    print(f"总计划数: {total_plans}")
    print(f"过滤掉非钢琴器乐: {skipped_instrumental} 条")
    print(f"\n方向分布:")
    for d, s in stats.items():
        print(f"  {d}: {s['count']} 条, 计划 {s['total_plan']} 人")
    print(f"\n批次分布: {dict(batch_counts)}")
    
    # 输出
    output = {
        'metadata': {
            'version': '2.0',
            'update_time': '2026-06-25',
            'source': '2026音乐生院校招生计划表（不完整）.xlsx',
            'year': 2026,
            'category': '艺术统考音乐类',
            'description': '2026年山东省音乐类招生计划（声乐+钢琴器乐+音乐教育）',
            'note': '器乐方向仅保留钢琴/键盘相关专业；数据来源于人工整理Excel'
        },
        'plans': plans,
        'statistics': {
            'total': len(plans),
            'schools': total_schools,
            'total_plan_count': total_plans,
            'directions': {k: v['count'] for k, v in stats.items()},
            'batches': dict(batch_counts),
        }
    }
    
    with open('plan_2026.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    size = os.path.getsize('plan_2026.json') / 1024
    print(f"\nplan_2026.json ({size:.1f} KB) 已生成")
    
    # 样本
    for d_name in ['vocal', 'instrumental', 'education']:
        samples = [p for p in plans if p['direction'] == d_name][:5]
        if samples:
            print(f"\n=== {d_name} ===")
            for r in samples:
                print(f"  {r['code']:6s} {r['name'][:14]:14s} | {r['major'][:50]:50s} | 计划={r['plan']} | {r['batch']}")

if __name__ == '__main__':
    main()
